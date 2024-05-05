import openpyxl
import uuid
import passlib.hash as hash
import traceback
import logging
import smtplib
from  email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def generate_unique_id(): # This function generates a unique identifier using the 'uuid' module
    return str(uuid.uuid4())

print("""
      *********************************************
              ** Welcome to Nairobi Restaurant **
      **********************************************""")

def get_user_info(): # This function prompts the user to enter their name and age, and validates the input 
    while True:
        name = input("What is your name? ")
        if len(name) < 3:
            print("Please enter a valid name with at least 3 characters.")
            continue
        age = input("What is your age? ")
        if validate_age(age) is None:
            print("Please enter a valid age.")
            continue
        return name, int(age)

def validate_age(age): # This function validates the user's age input
    try:
        age = int(age)
        if age > 0:
            return age
        else:
            return None
    except ValueError:
        return None
    
def hash_password(password):
    return hash.sha256_crypt.hash(password)
def verify_password(password, hashed_password):
    return hash.sha256_crypt.verify(password, hashed_password)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def log_info(message):
    logger.info(message)

def get_user_location(): # This function prompts the user to choose their loaction from a list of options
    location_options = {'a': 'Nairobi',
                        'b': 'Naivasha',
                        'c': 'Kisumu',
                        'd': 'Machakos',
                        'e': 'Nakuru',
                        'f': 'Eldoret',
                        'g': 'Kapsabet',
                        'h': 'Kakamega',
                        'i': 'Bungoma'}
    while True:
        print("Please choose your location: ")
        for key, value in location_options.items():
            print(f"{key}. {value}.")
        user_choice = input("\nEnter your choice (a-i) or 'q' to cancel:").lower()
        if user_choice in location_options:
            return location_options[user_choice]
        elif user_choice == 'q':
            return None
        else:
            print("Invalid choice. Please choose from the options you are given")

def get_user_order(): # This function prompts the user to choose their food order from a list of options
    order_options = {'a': 'Rice (KES 250)',
                     'b': 'Fish (KES 300)',
                     'c': 'Githeri (KES 150)',
                     'd': 'Salad (KES 100)',
                     'e': 'Chips (KES 120)',
                     'f': 'Meat (KES 200)',
                     'g': 'Noodles (KES 180)',
                     'h': 'Sushi (KES 400)'}
    while True:
        print("Please choose the food you wish to be delivered in your place")
        for key, value in order_options.items():
            print(f"{key}. {value}.")
        user_choice = input("\nEnter your food of choice (a-h) or 'q' to cancel: ").lower()
        if user_choice in order_options:
            return order_options[user_choice]
        elif user_choice == 'q':
            return None
        else:
            print("Invalid choice. Please choose from the options provided.")

def calculate_cost(order): # This function calculates the cost of the user's order based on the order name
    cost_options = {'Rice': 250,
                    'Fish': 300,
                    'Githeri': 150,
                    'Salad': 100,
                    'Chips': 120,
                    'Meat': 200,
                    'Noodles': 180,
                    'Sushi': 400}
    order_name = order.split()[0].replace('(', '').replace(')', '')
    if order_name in cost_options:
        return cost_options[order_name]
    else:
        print(f"Error: Invalid order name '{order_name}'")
        return None

def get_user_address(): # This function prompts the user to enter their address
    while True:
        address_parts = []
        house_number = input("Enter house number:")
        address_parts.append(house_number)
        street_name = input("Enter street name: ")
        address_parts.append(street_name)
        area_neighborhood = input("Enter area or neighborhood: ")
        address_parts.append(area_neighborhood)
        city = input("Enter city: ")
        address_parts.append(city)
        postal_zip_code = input("Enter postal/zip code: ")
        address_parts.append(postal_zip_code)
        address= ", ".join(address_parts) + "\n"
        return address
def confirm_order(name, age, location, order, address): # This function confirms the user's order and returns a boolean value indicating whether the order was confirmed
    print(f"\nSummary of your order:\nName: {name}\nAge: {age}\nLocation: {location}\nOrder: {order}\nAddress: {address}")
    while True:
        confirm = input("Please confirm your order by entering 'yes' or 'no': ").lower()
        if confirm == 'yes':
            return True
        elif confirm == 'no':
            return False
        else:
           print("Invalid response. Please enter 'yes' or 'no'.")


def save_to_excel(name, age, location, order, address, cost, identifier, rating): # This function saves the user's order to an Excel file
    # Open the Excel file
    try:
        workbook = openpyxl.load_workbook(r"D:\hospital\orders.xlsx")
        sheet_name = workbook.active.title
        sheet = workbook[sheet_name]
    except Exception as e:
        print(f"An error occured while saving to excel: {e}")
        traceback.print_exc() # print detailed error traceback for debugging

    # Find the last non-empty row
    max_row = sheet.max_row
    if max_row > 1:
        row = max_row + 1

    # Set column headers
    sheet.cell(row=1, column=1, value="Name")
    sheet.cell(row=1, column=2, value="Age")
    sheet.cell(row=1, column=3, value="Location")
    sheet.cell(row=1, column=4, value="Order")
    sheet.cell(row=1, column=5, value="Address")
    sheet.cell(row=1, column=6, value="Cost (KES)")
    sheet.cell(row=1, column=7, value="Identifier")
    

    # Set order details
    sheet.cell(row=row, column=1, value=name)
    sheet.cell(row=row, column=2, value=age)
    sheet.cell(row=row, column=3, value=location)
    sheet.cell(row=row, column=4, value=order)
    sheet.cell(row=row, column=5, value=address)
    sheet.cell(row=row, column=6, value=cost)
    sheet.cell(row=row, column=7, value=identifier)
    

    workbook.save(r"D:\hospital\orders.xlsx")
    return row + 1

# code for prompting user to rate their food order
def get_user_rating(): # This function prompts the user to rate their food order on a scale of 1 to 5 (1 being the lowest and 5 being the the highest)
    while True:
        rating = input("Please rate your food order on a scale of 1 to 5 (1 being the lowest and 5 being the the highest): ")
        if rating.isdigit() and int(rating) >= 1 and int(rating) <= 5:
            return int(rating)
        else:
            print("Please enter a valid rating.")

def send_email(name, age, location, order, address, cost, identifier, domain): # This function sends an email to the user with their order details 
    # Replace with your own email server credential
    sender_email = "gundumaxwell@gmail.com"
    receiver_email = f"{name}@{domain}"
    password = "ak47ak47"

    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = "Your Nairobi Restaurant Order"

    body = f"Dear {name}, \n\nThank you for your order Nairobi Restaurant!\n\nHere is a summary of your order:\nName: {name}\nAge: {age}\nLocation: {location}\nOrder: {order}\nAddress: {address}\nCost (KES): {cost}\nOrder Indentifier: {identifier}\n\nYour order will be delivered to your specified address. Thank you for choosing Nairobi Restaurant!\n\nBest regards,\nThe Nairobi Restaurant Team"

    message.attach(MIMEText(body, "plain"))

    # Use the SMTP library to send the email
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(sender_email, password)
        text = message.as_string()
        server.sendmail(sender_email, receiver_email, text)

def get_user_order_count(user_name): # This function return the number of orders placed by the user
    workbook = openpyxl.load_workbook(r"D:\hospital\orders.xlsx")
    sheet_name = workbook.active.title
    sheet = workbook[sheet_name]

    # Find the number of orders placed by the user
    order_count = 0
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == user_name:
            order_count += 1

    workbook.close()
    return order_count

if __name__ == "__main__":
    
    while True:
        name, age = get_user_info()
        if name is None or age is None:
            print("Order cancelled.")
            break
        else:
            location = get_user_location()

        if location is not None:
            order = get_user_order()

        if order is not None:
            address = get_user_address()

            # Calculate cost
            cost = calculate_cost(order)
            
            # Confirm order
            if confirm_order(name, age, location, order, address):
                print(f"\nYour order has been confirmed. The total cost is KES {cost}.")
                unique_id = generate_unique_id()
                
                # Prompt user to rate their food order
                rating = get_user_rating()


                # Save order details to excel workbook
                save_to_excel(name, age, location, order, address, cost, unique_id, rating)
                log_info("Order saved successfully to Excel Workbook.")

                # Offer a discount if the user has placed more than 5 orders
                order_count = get_user_order_count(name)
                if order_count > 5:
                    cost *= 0.9 # Apply a 10%  discount
                    print(f"Thank you for your loyalty! Your order total has been discounted by 10% and is now KES {cost}.")

                    # Send email
                    send_email(name, age, location, order, address, cost, unique_id)
                
            else:
                print("\nYour order has been cancelled.")
                continue
        else:
            print("\nYour order has been cancelled.")
            continue