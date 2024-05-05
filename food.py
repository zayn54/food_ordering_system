import openpyxl
import uuid
import passlib.hash as hash
import traceback
import logging
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import requests

def generate_unique_id(): # This function generates a unique identifier using the 'uuid' module
    return str(uuid.uuid4())

print("""
      *******************************************
      ** Welcome to the Food Ordering System **
      *******************************************
      """)

wb = openpyxl.load_workbook('menu.xlsx') # This line loads the menu from an Excel file
sheet = wb.active # This line selects the active sheet in the workbook

print("Here is the menu:")

for row in range(2, sheet.max_row + 1): # This loop prints the menu
    print(f"{sheet.cell(row=row, column=1).value} - KES {sheet.cell(row=row, column=2).value}")

order_id = generate_unique_id() # This line generates a unique identifier for the order

print(f"\nYour order id is {order_id}.")

order = {} # This line initializes an empty dictionary to store the order

while True: # This loop allows the user to add items to their order
    item_code = input("Enter the item code of the item you want to order: ")
    quantity = int(input("Enter the quantity of the item: "))

    if item_code not in [item.value for item in sheet['A']]: # This checks if the item code is valid
        print("Invalid item code. Please try again.")
        continue

    if quantity <= 0: # This checks if the quantity is valid
        print("Invalid quantity. Please try again.")
        continue

    if item_code in order: # This checks if the item is already in the order
        order[item_code] += quantity # If it is, the quantity is updated
    else: # If it is not, the item is added to the order
        order[item_code] = quantity

    add_more = input("Do you want to add more items to your order? (yes/no): ")
    if add_more.lower()!= 'yes':
        break

total_price = sum([sheet.cell(row=sheet.cell(column=1, row=item_code).row, column=2).value * quantity for item_code, quantity in order.items()]) # Calculates the total price of the order

print(f"\nYour order:")
for item, price in order.items():
    print(f"{item}: KES {price}")

confirm_order = input("Is this correct? (yes/no): ")

if confirm_order.lower()!= "yes":
    print("Okay, please come back when you are ready to order.")
    exit()

customer_name = input("What is your name? ")
customer_email = input("What is your email address? ")
customer_phone = input("What is your phone number? ")

order_id = generate_unique_id() # Generates a unique identifier for the order

print(f"\nYour order has been received and will be processed with the order ID {order_id}.")

# Saves the order details to an Excel file
wb_orders = openpyxl.load_workbook('orders.xlsx')
sheet_orders = wb_orders.active
sheet_orders.append((order_id, customer_name, customer_email, customer_phone, total_price, *[item for item, price in order.items()]))
wb_orders.save('orders.xlsx')

# Sends an email to the customer with the order details
sender_email = "you@example.com"
sender_password = "your_password"
recipient_email = customer_email

message = MIMEMultipart()
message['From'] = sender_email
message['To'] = recipient_email
message['Subject'] = "Order Confirmation"

message.attach(MIMEText(f"Dear {customer_name},\n\nThank you for your order. Your order details are as follows:\n\nOrder ID: {order_id}\nCustomer Name: {customer_name}\nCustomer Email: {customer_email}\nCustomer Phone: {customer_phone}\nTotal Price: KES {total_price}\nOrder Items: {order}\n\nYour order will be processed and delivered as soon as possible.\n\nBest regards,\nThe Food Ordering System Team"))

try:
    server = smtplib.SMTP('smtp.example.com', 587)
    server.starttls()
    server.login(sender_email, sender_password)
    text = message.as_string()
    server.sendmail(sender_email, recipient_email, text)
    server.quit()
    print("Email sent!")
except Exception as e:
    print("Error sending email:", traceback.format_exc())

# Allows the customer to pay for their order using M-Pesa
paybill_number = "123456" # Replace with your M-Pesa paybill number
account_number = order_id # Replace with your M-Pesa account number (optional)
amount = total_price